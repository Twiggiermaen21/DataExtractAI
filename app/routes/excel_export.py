import os
import re
import json
import logging
from io import BytesIO

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import Blueprint, request, jsonify, current_app, send_file

excel_export_bp = Blueprint('excel_export', __name__)
log = logging.getLogger(__name__)

# KolejnoĹ›Ä‡ i etykiety wszystkich moĹĽliwych kolumn
COLUMN_DEFS = [
    ("numer_faktury",            "Nr faktury",                  False),
    ("sprzedawca",               "Sprzedawca",                  False),
    ("data_wystawienia",         "Data wystawienia",             False),
    ("data_sprzedazy",           "Data sprzedaży",              False),
    ("wolumen_energii",          "Wolumen energii [kWh]",        True),
    ("kwota_netto",              "Kwota netto [zł]",             True),
    ("kwota_brutto",             "Kwota brutto [zł]",            True),
    ("razem",                    "Razem do zapłaty [zł]",        True),
    ("kwota_vat",                "Kwota VAT [zł]",               True),
    ("sprzedaz_cena_netto",      "Sprzedaż netto [zł]",         True),
    ("sprzedaz_cena_brutto",     "Sprzedaż brutto [zł]",        True),
    ("dystrybucja_cena_netto",   "Dystrybucja netto [zł]",      True),
    ("dystrybucja_cena_brutto",  "Dystrybucja brutto [zł]",     True),
    ("oplata_abonamentowa",             "Abonamentowa netto [zł]",       True),
    ("oplata_abonamentowa_brutto",      "Abonamentowa brutto [zł]",      True),
    ("oplata_sieciowa_stala",           "Sieciowa stała netto [zł]",     True),
    ("oplata_sieciowa_stala_brutto",    "Sieciowa stała brutto [zł]",    True),
    ("oplata_sieciowa_zmienna",         "Sieciowa zmienna netto [zł]",   True),
    ("oplata_sieciowa_zmienna_brutto",  "Sieciowa zmienna brutto [zł]",  True),
    ("oplata_jakosciowa",               "Jakościowa netto [zł]",         True),
    ("oplata_jakosciowa_brutto",        "Jakościowa brutto [zł]",        True),
    ("oplata_oze",                      "OZE netto [zł]",                True),
    ("oplata_oze_brutto",               "OZE brutto [zł]",               True),
    ("oplata_kogeneracyjna",            "Kogeneracyjna netto [zł]",      True),
    ("oplata_kogeneracyjna_brutto",     "Kogeneracyjna brutto [zł]",     True),
    ("oplata_przejsciowa",              "Przejściowa netto [zł]",        True),
    ("oplata_przejsciowa_brutto",       "Przejściowa brutto [zł]",       True),
    ("oplata_mocowa",                   "Mocowa netto [zł]",             True),
    ("oplata_mocowa_brutto",            "Mocowa brutto [zł]",            True),
    ("naleznos_netto",           "Należność netto [zł]",        True),
    ("naleznos_brutto",          "Należność brutto [zł]",       True),
]

# Indeksy (key â†’ (label, is_numeric))
COLUMN_MAP = {k: (label, numeric) for k, label, numeric in COLUMN_DEFS}

# Pary netto/brutto â†’ scalone nagĹ‚Ăłwki grupujÄ…ce w Excelu
# klucz: id_netto, wartoĹ›Ä‡: (etykieta grupy, id_brutto)
COLUMN_PAIRS = {
    "naleznos_netto":         ("Należność",         "naleznos_brutto"),
    "kwota_netto":            ("Kwoty",              "kwota_brutto"),
    "sprzedaz_cena_netto":    ("Sprzedaż energii",   "sprzedaz_cena_brutto"),
    "dystrybucja_cena_netto": ("Dystrybucja",        "dystrybucja_cena_brutto"),
    "oplata_abonamentowa":    ("Abonamentowa",       "oplata_abonamentowa_brutto"),
    "oplata_sieciowa_stala":  ("Sieciowa stała",     "oplata_sieciowa_stala_brutto"),
    "oplata_sieciowa_zmienna":("Sieciowa zmienna",   "oplata_sieciowa_zmienna_brutto"),
    "oplata_jakosciowa":      ("Jakościowa",         "oplata_jakosciowa_brutto"),
    "oplata_oze":             ("OZE",                "oplata_oze_brutto"),
    "oplata_kogeneracyjna":   ("Kogeneracyjna",      "oplata_kogeneracyjna_brutto"),
    "oplata_przejsciowa":     ("Przejściowa",        "oplata_przejsciowa_brutto"),
    "oplata_mocowa":          ("Mocowa",             "oplata_mocowa_brutto"),
}


def _to_number(value):
    """Zwraca float z wartoĹ›ci ktĂłra moĹĽe zawieraÄ‡ jednostki (kWh, zĹ‚ itp.)."""
    if value is None:
        return None
    s = str(value).strip()
    # UsuĹ„ jednostki na koĹ„cu
    s = re.sub(r'[\s\u00a0]*[a-zA-ZĹ‚zĹ‚ĹZĹ%]+[\s\u00a0]*$', '', s).strip()
    # UsuĹ„ spacje i &nbsp; wewnÄ…trz liczby
    s = s.replace('\u00a0', '').replace(' ', '')
    # Polski format: 1.234,56 â†’ 1234.56
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _find_json(output_dir, filename):
    """Szuka pliku JSON w OUTPUT_FOLDER i podfolderze wezwania_faktury."""
    for candidate in [
        os.path.join(output_dir, filename),
        os.path.join(output_dir, 'wezwania_faktury', filename),
    ]:
        if os.path.exists(candidate):
            return candidate

    # Fallback: szukaj po nazwie bazowej
    base = os.path.splitext(filename)[0]
    for f in os.listdir(output_dir):
        if f.endswith('.json') and base in f:
            return os.path.join(output_dir, f)
    return None


@excel_export_bp.route('/api/export_excel', methods=['POST'])
def export_excel():
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'Brak danych do eksportu'}), 400

    files = data.get('files') or []
    documents = data.get('documents') or []
    if not files and not documents:
        return jsonify({'success': False, 'error': 'Brak danych do eksportu'}), 400

    # Kolumny wybrane przez użytkownika (opcjonalne — jeśli brak, bierzemy wszystkie)
    selected = data.get('selected_columns') or [k for k, _, _ in COLUMN_DEFS]
    output_dir = current_app.config['OUTPUT_FOLDER']

    # Wyznacz aktywne kolumny zachowując kolejność z COLUMN_DEFS
    active_cols = [(k, label, numeric)
                   for k, label, numeric in COLUMN_DEFS
                   if k in selected]

    if not active_cols:
        return jsonify({'success': False, 'error': 'Nie wybrano żadnych kolumn'}), 400

    records = []  # lista (dict_of_values, is_vision, error_msg_or_None)

    if documents:
        for doc in documents:
            fields = doc.get('fields') or {}
            if isinstance(fields, str):
                try:
                    fields = json.loads(fields)
                except Exception:
                    fields = {}

            parse_error = fields.get('_parse_error') if isinstance(fields, dict) else None
            is_vision = bool(doc.get('is_vision', False))
            source = doc.get('filename') or doc.get('source_file') or 'dokument'

            row = {'_source': source}
            for col_id, _, is_numeric in active_cols:
                raw = fields.get(col_id)
                if raw is None:
                    row[col_id] = None
                elif is_numeric and '|' in str(raw):
                    row[col_id] = str(raw)
                elif is_numeric:
                    row[col_id] = _to_number(raw)
                else:
                    row[col_id] = str(raw) if raw is not None else ''

            records.append((row, is_vision, parse_error))
    else:
        for filename in files:
            json_path = _find_json(output_dir, filename)
            if not json_path:
                log.warning("Nie znaleziono JSON dla: %s", filename)
                records.append(({'_source': filename}, False, f'Brak pliku wyników (JSON) dla: {filename}'))
                continue
            try:
                with open(json_path, 'r', encoding='utf-8') as f:
                    doc = json.load(f)

                fields = doc.get('extracted_fields') or doc.get('extracted_data') or doc.get('fields') or {}
                if isinstance(fields, str):
                    try:
                        fields = json.loads(fields)
                    except Exception:
                        fields = {}

                parse_error = fields.get('_parse_error') if isinstance(fields, dict) else None
                is_vision = doc.get('is_vision', False)
                source = doc.get('source_file', filename)

                row = {'_source': source}
                for col_id, _, is_numeric in active_cols:
                    raw = fields.get(col_id)
                    if raw is None:
                        row[col_id] = None
                    elif is_numeric and '|' in str(raw):
                        row[col_id] = str(raw)
                    elif is_numeric:
                        row[col_id] = _to_number(raw)
                    else:
                        row[col_id] = str(raw) if raw is not None else ''

                records.append((row, is_vision, parse_error))
            except Exception as exc:
                log.exception("Błąd odczytu JSON: %s", json_path)
                records.append(({'_source': filename}, False, str(exc)))

    if not records:
        return jsonify({'success': False, 'error': 'Nie znaleziono poprawnych danych do eksportu'}), 404

    # â”€â”€ Budowanie pliku Excel â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Faktury Energia'

    # Style
    header_fill = PatternFill(fill_type='solid', fgColor='1F2937')   # ciemny nagĹ‚Ăłwek
    header_font = Font(bold=True, color='FFFFFF', size=10)
    scan_fill   = PatternFill(fill_type='solid', fgColor='FFF9C4')   # ĹĽĂłĹ‚ty dla skanĂłw
    alt_fill    = PatternFill(fill_type='solid', fgColor='F3F4F6')   # szary co 2. rekord
    error_fill  = PatternFill(fill_type='solid', fgColor='FEE2E2')   # czerwony dla bĹ‚Ä™dĂłw
    sum_fill    = PatternFill(fill_type='solid', fgColor='D1FAE5')   # zielony dla SUM
    sum_font    = Font(bold=True, size=10)
    center      = Alignment(horizontal='center', vertical='center')
    thin        = Side(style='thin', color='D1D5DB')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    # â”€â”€ Grupowanie kolumn (pary netto/brutto â†’ scalony nagĹ‚Ăłwek) â”€â”€â”€â”€
    active_ids = {k for k, _, _ in active_cols}

    # Buduj listÄ™ grup zachowujÄ…c kolejnoĹ›Ä‡ COLUMN_DEFS; pomijaj brutto ktĂłre juĹĽ
    # sÄ… wchĹ‚oniÄ™te przez odpowiedniÄ… grupÄ™ netto
    brutto_absorbed = set()
    header_groups = []
    for col_id, label, is_numeric in active_cols:
        if col_id in brutto_absorbed:
            continue
        if col_id in COLUMN_PAIRS and COLUMN_PAIRS[col_id][1] in active_ids:
            group_label, brutto_id = COLUMN_PAIRS[col_id]
            brutto_absorbed.add(brutto_id)
            header_groups.append({
                'label': group_label,
                'cols': [(col_id, 'Netto', is_numeric), (brutto_id, 'Brutto', True)],
                'paired': True,
            })
        else:
            header_groups.append({
                'label': label,
                'cols': [(col_id, None, is_numeric)],
                'paired': False,
            })

    has_pairs = any(g['paired'] for g in header_groups)
    data_row_start = 3 if has_pairs else 2
    total_data_cols = sum(len(g['cols']) for g in header_groups)
    skan_col = total_data_cols + 2  # +1 dla "Plik źródłowy", +1 za skan

    sub_fill   = PatternFill(fill_type='solid', fgColor='374151')  # nieco jaĹ›niejszy nagĹ‚Ăłwek sub
    sub_font   = Font(bold=True, color='D1D5DB', size=9)

    # â”€â”€ NagĹ‚Ăłwek â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    if has_pairs:
        ws.row_dimensions[1].height = 20
        ws.row_dimensions[2].height = 16

        # KomĂłrka "Plik źródłowy" â€” scalona pionowo (wiersze 1â€“2)
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        c = ws.cell(row=1, column=1, value='Plik źródłowy')
        c.fill = header_fill; c.font = header_font; c.alignment = center; c.border = border

        col_pos = 2
        for g in header_groups:
            if g['paired']:
                end_pos = col_pos + len(g['cols']) - 1
                # Wiersz 1 â€” scalona etykieta grupy
                ws.merge_cells(start_row=1, start_column=col_pos, end_row=1, end_column=end_pos)
                c = ws.cell(row=1, column=col_pos, value=g['label'])
                c.fill = header_fill; c.font = header_font
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = border
                # Wiersz 2 â€” Netto / Brutto
                for i, (_, sub_label, _) in enumerate(g['cols']):
                    c2 = ws.cell(row=2, column=col_pos + i, value=sub_label)
                    c2.fill = sub_fill; c2.font = sub_font
                    c2.alignment = center; c2.border = border
                col_pos += len(g['cols'])
            else:
                # Pojedyncza kolumna â€” scalona pionowo (wiersze 1â€“2)
                ws.merge_cells(start_row=1, start_column=col_pos, end_row=2, end_column=col_pos)
                c = ws.cell(row=1, column=col_pos, value=g['label'])
                c.fill = header_fill; c.font = header_font; c.alignment = center; c.border = border
                col_pos += 1

        # "Skan?" â€” scalona pionowo
        ws.merge_cells(start_row=1, start_column=skan_col, end_row=2, end_column=skan_col)
        c = ws.cell(row=1, column=skan_col, value='Skan?')
        c.fill = header_fill; c.font = header_font; c.alignment = center; c.border = border
    else:
        col_pos = 2
        ws.cell(row=1, column=1, value='Plik źródłowy')
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).font = header_font
        ws.cell(row=1, column=1).alignment = center
        ws.cell(row=1, column=1).border = border
        for g in header_groups:
            c = ws.cell(row=1, column=col_pos, value=g['label'])
            c.fill = header_fill; c.font = header_font; c.alignment = center; c.border = border
            col_pos += 1
        ws.cell(row=1, column=skan_col, value='Skan?').fill = header_fill
        ws.cell(row=1, column=skan_col).font = header_font
        ws.cell(row=1, column=skan_col).alignment = center
        ws.cell(row=1, column=skan_col).border = border
        ws.row_dimensions[1].height = 22

    # â”€â”€ Dane â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # Buduj mapÄ™ col_id â†’ pozycja kolumny (uwzglÄ™dnia pary)
    col_position = {}
    pos = 2
    for g in header_groups:
        for col_id, _, _ in g['cols']:
            col_position[col_id] = pos
            pos += 1

    def _split_multi(val):
        """Zwraca listÄ™ wartoĹ›ci (float lub None). JeĹ›li brak "|" â†’ lista jednoelementowa."""
        if isinstance(val, str) and '|' in val:
            return [_to_number(v.strip()) for v in val.split('|') if v.strip()]
        return [val]

    current_row = data_row_start

    for rec_idx, (row, is_vision, parse_error) in enumerate(records):
        # Wiersz bĹ‚Ä™du â€” czerwone tĹ‚o, komunikat przez caĹ‚Ä… szerokoĹ›Ä‡
        if parse_error:
            err_font = Font(bold=True, color='991B1B', size=9)
            c = ws.cell(row=current_row, column=1, value=f'BŁĄD - {row.get("_source", "?")}')
            c.fill = error_fill; c.font = err_font; c.border = border
            c2 = ws.cell(row=current_row, column=2, value=f'Błąd parsowania odpowiedzi LLM: {parse_error}')
            c2.fill = error_fill; c2.font = Font(color='991B1B', size=8, italic=True); c2.border = border
            if skan_col > 2:
                ws.merge_cells(start_row=current_row, start_column=2,
                               end_row=current_row, end_column=skan_col)
            current_row += 1
            continue

        fill = scan_fill if is_vision else (alt_fill if rec_idx % 2 == 1 else None)

        # Ustal ile pod-wierszy potrzeba (max liczba wartoĹ›ci w kolumnach multi)
        sub_count = max(
            (len(_split_multi(row.get(col_id))) for col_id, _, _ in active_cols),
            default=1
        )
        end_row = current_row + sub_count - 1

        def _write_cell(r, col, value, is_num=False, align_center=False):
            c = ws.cell(row=r, column=col, value=value)
            if align_center:
                c.alignment = center
            else:
                c.alignment = Alignment(
                    horizontal='right' if is_num else 'left', vertical='center'
                )
            if is_num and value is not None:
                c.number_format = '#,##0.00'
            c.border = border
            if fill:
                c.fill = fill
            return c

        def _merge_if_multi(col):
            """Scal komĂłrki pionowo gdy rekord zajmuje kilka wierszy."""
            if sub_count > 1:
                ws.merge_cells(
                    start_row=current_row, start_column=col,
                    end_row=end_row, end_column=col
                )
                c = ws.cell(row=current_row, column=col)
                c.alignment = Alignment(vertical='center', wrap_text=False)

        # Kolumna "Plik źródłowy" â€” scalona
        _write_cell(current_row, 1, row.get('_source', ''))
        _merge_if_multi(1)

        for col_id, _, is_numeric in active_cols:
            col_offset = col_position.get(col_id)
            if col_offset is None:
                continue
            vals = _split_multi(row.get(col_id))

            if len(vals) > 1:
                # KaĹĽda wartoĹ›Ä‡ w osobnym wierszu â€” prawdziwe liczby
                for sub_i, v in enumerate(vals):
                    num = v if isinstance(v, float) else _to_number(v)
                    _write_cell(current_row + sub_i, col_offset, num, is_num=is_numeric)
                # Puste komĂłrki w wierszach gdzie brak wartoĹ›ci
                for sub_i in range(len(vals), sub_count):
                    _write_cell(current_row + sub_i, col_offset, None, is_num=is_numeric)
            else:
                val = vals[0]
                _write_cell(current_row, col_offset, val, is_num=is_numeric)
                _merge_if_multi(col_offset)

        # Kolumna "Skan?" â€” scalona
        _write_cell(current_row, skan_col, 'TAK' if is_vision else '', align_center=True)
        _merge_if_multi(skan_col)

        current_row += sub_count

    # â”€â”€ Wiersz SUM â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    sum_row = current_row

    ws.cell(row=sum_row, column=1, value='SUMA').font = sum_font

    for col_id, _, is_numeric in active_cols:
        col_offset = col_position.get(col_id)
        if col_offset is None:
            continue
        c = ws.cell(row=sum_row, column=col_offset)
        if is_numeric:
            col_letter = get_column_letter(col_offset)
            c.value = f'=SUM({col_letter}{data_row_start}:{col_letter}{sum_row - 1})'
            c.number_format = '#,##0.00'
            c.font = sum_font
        c.fill = sum_fill
        c.border = border
        c.alignment = Alignment(horizontal='right', vertical='center')

    ws.cell(row=sum_row, column=skan_col).fill = sum_fill

    # â”€â”€ OstrzeĹĽenie o skanach â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    scan_count = sum(1 for _, is_vision, _ in records if is_vision)
    if scan_count > 0:
        note_row = sum_row + 2
        note = ws.cell(
            row=note_row, column=1,
            value=f'UWAGA: {scan_count} plik(ów) to skany (żółte wiersze) - wyższe ryzyko błędu OCR. Zweryfikuj ręcznie.'
        )
        note.font = Font(color='B45309', italic=True, size=9)
        ws.merge_cells(
            start_row=note_row, start_column=1,
            end_row=note_row, end_column=skan_col
        )

    # â”€â”€ Autodopasowanie szerokoĹ›ci â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    for col_idx in range(1, skan_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or ''))
             for r in range(1, sum_row)),
            default=5
        )
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    ws.freeze_panes = f'A{data_row_start}'

    # â”€â”€ Zapis do pamiÄ™ci i wysyĹ‚ka â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='raport_faktury_energia.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


